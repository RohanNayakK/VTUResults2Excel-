#package import
import openpyxl
import json
import sys

#get path to write generated file
filePath = sys.argv[1]

# Opening JSON file
f = open('data.json')

# returns JSON object as a dictionary
data = json.load(f)

# Closing file
f.close()

#open excel workbooka and load sheet
wb_obj = openpyxl.load_workbook('newTemplate.xlsx')
sheet_obj = wb_obj["Sheet1"]

#initial values of row/col
row=5
col=1

subIaCol=-1
subEaCol=0
subTotalCol=1
subResCol=2

subnameCol=3
subCodeCol=3

#flag to check first iteration
firstTime=True

for i in data:
    #enter subject header details (subcode, subname)
    if(firstTime):
        for k in range(len(i['subjects'])-1):
            if(k == 0):
                sheet_obj['C2']= i['subjects'][0]['subjectName']
                sheet_obj['C3']= i['subjects'][0]['subjectCode']

            if(k == 1):
                sheet_obj['G2']= i['subjects'][1]['subjectName']
                sheet_obj['G3']= i['subjects'][1]['subjectCode']

            if(k == 2):
                sheet_obj['K2']= i['subjects'][2]['subjectName']
                sheet_obj['K3']= i['subjects'][2]['subjectCode']

            if(k == 3):
                sheet_obj['O2']= i['subjects'][3]['subjectName']
                sheet_obj['O3']= i['subjects'][3]['subjectCode']

            if(k == 4):
                sheet_obj['S2']= i['subjects'][4]['subjectName']
                sheet_obj['S3']= i['subjects'][4]['subjectCode']

            if(k == 4):
                sheet_obj['W2']= i['subjects'][5]['subjectName']
                sheet_obj['W3']= i['subjects'][5]['subjectCode']

            if(k == 5):
                sheet_obj['AA2']= i['subjects'][6]['subjectName']
                sheet_obj['AA3']= i['subjects'][6]['subjectCode']

            if(k == 6):
                sheet_obj['AE2']= i['subjects'][7]['subjectName']
                sheet_obj['AE3']= i['subjects'][7]['subjectCode']

        firstTime=False


    #enter name and USN
    nameCell = sheet_obj.cell(row=row, column=1)
    usnCell = sheet_obj.cell(row=row, column=2)

    #enter each subject mark in a row
    for j in range(len(i['subjects'])):
        if(i['subjects'][j]):
            subIaCol=subIaCol+4
            subEaCol=subEaCol+4
            subTotalCol=subTotalCol+4
            subResCol=subResCol+4
            subIA=sheet_obj.cell(row=row,column=subIaCol)
            subEA=sheet_obj.cell(row=row,column=subEaCol)
            subTotal=sheet_obj.cell(row=row,column=subTotalCol)
            subResult=sheet_obj.cell(row=row,column=subResCol)
            nameCell.value=i['studentName']
            usnCell.value=i['studentUSN']
            subIA.value=i['subjects'][j]['iaMarks']
            subEA.value=i['subjects'][j]['eaMarks']
            subTotal.value=i['subjects'][j]['totalMarks']
            subResult.value=i['subjects'][j]['result']

    #change of row
    subIaCol=-1
    subEaCol=0
    subTotalCol=1
    subResCol=2
    row+=1

#save file
 




